from fastapi import FastAPI
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import httpx
import base64
import json
import os
import io
import time
from openpyxl import load_workbook

app = FastAPI()

GITHUB_API = "https://api.github.com/repos/openfootball/worldcup.json/contents/2026/worldcup.json"
GITHUB_RAW = "https://raw.githubusercontent.com/openfootball/worldcup.json/master/2026/worldcup.json"

# Yellow/red cards are maintained by hand in "Card stats.xlsx" at the repo root.
# /api/cards fetches that file from GitHub's raw URL on each request (cached for
# CARD_CACHE_TTL seconds) so editing it on GitHub goes live WITHOUT a redeploy.
# The copy committed alongside the app is used as a local fallback if the fetch
# fails. Columns: Record, Team1, YC_Team1, RC_Team1, Team2, YC_Team2, RC_Team2.
CARD_RAW_URL = "https://raw.githubusercontent.com/VladV2021/football-caliber/main/Card%20stats.xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))
CARD_FILE = os.path.join(HERE, "Card stats.xlsx")
CARD_CACHE_TTL = 60  # seconds
_card_cache = {"ts": 0.0, "data": None}


@app.get("/api/matches")
async def get_matches():
    """Proxy the openfootball GitHub feed server-side — no CORS issues.
    Tries the Contents API first (structured JSON), falls back to raw URL."""
    async with httpx.AsyncClient(timeout=10) as client:
        # Attempt 1: GitHub Contents API (base64-encoded JSON)
        try:
            r = await client.get(GITHUB_API, headers={"Accept": "application/vnd.github.v3+json"})
            if r.status_code == 200:
                meta = r.json()
                if "content" in meta:
                    raw = base64.b64decode(meta["content"].replace("\n", "")).decode("utf-8")
                    data = json.loads(raw)
                    return {"ok": True, "source": "github-api", "matches": data.get("matches", [])}
        except Exception:
            pass

        # Attempt 2: raw.githubusercontent.com
        try:
            r = await client.get(GITHUB_RAW)
            if r.status_code == 200:
                data = r.json()
                return {"ok": True, "source": "github-raw", "matches": data.get("matches", [])}
        except Exception:
            pass

    return {"ok": False, "error": "Both GitHub sources unavailable", "matches": []}


def _num(v):
    """Coerce a spreadsheet cell to a non-negative int (blank/garbage -> 0)."""
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return 0


def _parse_card_xlsx(source) -> dict:
    """Parse 'Card stats.xlsx' (a path or a file-like object) into
    {"Team1|Team2": {yc1,rc1,yc2,rc2}}. Columns are matched by header name,
    so column order can change freely."""
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return {}
        header = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}

        needed = ["Team1", "YC_Team1", "RC_Team1", "Team2", "YC_Team2", "RC_Team2"]
        missing = [k for k in needed if k not in header]
        if missing:
            raise ValueError(f"Spreadsheet missing columns {missing}. Found: {list(header)}")

        def cell(row, key):
            i = header[key]
            return row[i] if i < len(row) else None

        out = {}
        for row in rows:
            if row is None:
                continue
            t1 = cell(row, "Team1")
            t2 = cell(row, "Team2")
            t1 = str(t1).strip() if t1 is not None else ""
            t2 = str(t2).strip() if t2 is not None else ""
            if not t1 or not t2:
                continue
            out[f"{t1}|{t2}"] = {
                "yc1": _num(cell(row, "YC_Team1")),
                "rc1": _num(cell(row, "RC_Team1")),
                "yc2": _num(cell(row, "YC_Team2")),
                "rc2": _num(cell(row, "RC_Team2")),
            }
        return out
    finally:
        wb.close()


@app.get("/api/cards")
async def get_cards():
    """Serve manually-tracked yellow/red card data. Fetches 'Card stats.xlsx'
    from GitHub (cached CARD_CACHE_TTL s) so edits go live with no redeploy;
    falls back to the committed copy, then to stale cache."""
    now = time.time()
    if _card_cache["data"] is not None and (now - _card_cache["ts"]) < CARD_CACHE_TTL:
        return {"ok": True, "source": "cache", "cards": _card_cache["data"],
                "count": len(_card_cache["data"])}

    # Primary: live file from GitHub raw
    try:
        async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
            r = await client.get(CARD_RAW_URL, headers={"User-Agent": "football-caliber/1.0"})
        if r.status_code == 200 and r.content[:2] == b"PK":
            cards = _parse_card_xlsx(io.BytesIO(r.content))
            _card_cache["data"] = cards
            _card_cache["ts"] = now
            return {"ok": True, "source": "github-raw", "cards": cards, "count": len(cards)}
        raise ValueError(f"Unexpected response (status {r.status_code})")
    except Exception as live_err:
        # Fallback: the copy committed next to the app
        try:
            cards = _parse_card_xlsx(CARD_FILE)
            _card_cache["data"] = cards
            _card_cache["ts"] = now
            return {"ok": True, "source": "repo-file-fallback", "cards": cards,
                    "count": len(cards), "warning": str(live_err)}
        except Exception as file_err:
            if _card_cache["data"] is not None:
                return {"ok": True, "source": "stale-cache", "cards": _card_cache["data"],
                        "count": len(_card_cache["data"]), "warning": str(file_err)}
            return {"ok": False, "error": str(file_err), "cards": {}}


@app.get("/health")
async def health():
    return {"status": "ok"}

# Serve the frontend
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def root():
    return FileResponse("static/index.html")
